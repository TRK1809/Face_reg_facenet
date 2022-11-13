import keras_facenet
import numpy as np
from numpy import asarray
import cv2
import os
import pickle
import PIL
from PIL import Image as Img
from keras_facenet import FaceNet
import time
from openpyxl import Workbook, load_workbook
loaded_model = pickle.load(open('finalized_model.sav', 'rb'))
def pushtoxl(event,day,month,year):
    wb = load_workbook("C:\\Users\\Admin\\PycharmProjects\\FACE_REG\\Diem_Danh.xlsx")
    sheetname = "Time_{}_{}_{}".format(day, month, year)
    print(sheetname)
    if sheetname in wb.sheetnames:
        sheet = wb[sheetname]
        sheet = wb.active
        print("sheet exist")


        c1 = sheet.cell(row=sheet.max_row + 1, column=1)
        c1.value = identity
        c2 = sheet.cell(row=c1.row, column=2)
        c2.value = str(clock)
        c3 = sheet.cell(row=c1.row, column=3)
        c3.value = event
    else:
        ws2 = wb.create_sheet("{}".format(sheetname))

        ws2['A1'] = "Names"
        ws2['B1'] = "Time"
        ws2['C1'] = "Event"
        c1 = ws2.cell(row=ws2.max_row + 1, column=1)
        c1.value = identity
        c2 = ws2.cell(row=c1.row, column=2)
        c2.value = str(clock)
        c3 = ws2.cell(row=c1.row, column=3)
        c3.value = event
    wb.save("C:\\Users\\Admin\\PycharmProjects\\FACE_REG\\Diem_Danh.xlsx")

    return

def check_in():
    print('{} check in at {}'.format(identity,clock))
    print("Have a Nice Day")
    event ="check in"
    pushtoxl(event,day,month,year)
    return
def check_out():
    print('{} check out at {}'.format(identity, clock))
    print("Have a Nice Day")
    event = "check out"
    pushtoxl(event,day,month,year)
    return

# #read file vector
# myfile = open('new_data.pkl','rb')
# data = pickle.load(myfile)
# myfile.close()
#init WEbcam
video_cap = cv2.VideoCapture(0)
# read Haarcascade file to face detected
cascPath=os.path.dirname(cv2.__file__)+"/data/haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascPath)
embedder = FaceNet()
while 1:
    #clock init
    clock = time.localtime()


    day = clock.tm_mday
    month = clock .tm_mon
    year = clock.tm_year
    clock = time.asctime(clock)

    ret,frames = video_cap.read()
    gray = cv2.cvtColor(frames,cv2.COLOR_BGR2GRAY)
    imgcopy = frames.copy()
    faces = faceCascade.detectMultiScale(
        gray,
        scaleFactor=1.1,
        minNeighbors=5,
        minSize=(30, 30),
        flags=cv2.CASCADE_SCALE_IMAGE
    ) # detect face
    for (x, y, w, h) in faces:
        coords = [x, y, w, h]
        cv2.rectangle(frames, (x, y), (x + w, y + h), (0, 255, 0), 2)
        crop = imgcopy[coords[1]: coords[1] + coords[3], coords[0]: coords[0] + coords[2]]

        crop = cv2.cvtColor(crop, cv2.COLOR_BGR2RGB)
        crop = Img.fromarray(crop)
        crop = asarray(crop)
        crop = crop.astype('float32')
        crop = cv2.resize(crop, (160, 160))
        crop = crop.reshape(1, 160, 160, 3)
        signature = embedder.embeddings(crop)
        identity = str(loaded_model.predict(signature)[0])

        # min_dist = 10
        # identity = ''
        # for key, value in data.items():
        #     dist = np.linalg.norm(value - signature)
        #     if dist < min_dist:
        #         min_dist = dist
        #         identity = key
        cv2.putText(frames,identity,(x, y - 4),cv2.FONT_HERSHEY_SIMPLEX,0.8,(255,255,0,),1,cv2.LINE_AA)
        cv2.putText(frames,clock,(x, y+h),cv2.FONT_HERSHEY_SIMPLEX,0.8,(255,0,255),1,cv2.LINE_AA)
    cv2.imshow('video',frames)
    key = cv2.waitKey(8) &0xFF
    if key == ord('q'):
        break
    if key == ord('s'):
        check_in()
    if key == ord('c'):
        check_out()

video_cap.release()
cv2.destroyAllWindows()


print("Have a Nice Day")